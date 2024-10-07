from tkinter import *
from tkinter import ttk
import datetime
import openpyxl
from openpyxl import Workbook 
class win:
    def __init__ (self,wid_long,title,ico):
        self.wid_long=wid_long
        self.title=title
        self.ico=ico
        win=Tk()
        now=datetime.datetime.now()
        date=now.strftime("%y-%m-%d")
        ##########exel############
        wb=Workbook()
        ws=wb.active
        ws.title='customersbill'
        ws["A1"]='name'
        ws["B1"]='total'
        ws["C1"]='date'
        
        wb.save('customersbill.xlsx')

        ####################     
        dict_price={0:['buton1',10],1:['buton2',20]}
        ########insert to exels file
        def insertxlsx():
            excel=openpyxl.load_workbook('customersbill.xlsx')
            file=excel.active
            
            file.cell(column=1,row=file.max_row+1,value=en_name.get())
            file.cell(column=2,row=file.max_row,value=en_total.get())
            file.cell(column=3,row=file.max_row,value=en_date.get())
            excel.save('customersbill.xlsx')

        ##########
        def bill():
           global en_name
           global en_date
           global en_total
           win.geometry('1170x552')
           customer_frame=Frame(win,bg='silver',width=240,height=350)
           customer_frame.place(x=920,y=1)
           l_name=Label(customer_frame,text="اسم المشتري",bg='silver',fg='white')
           l_name.place(x=168,y=10)
           en_name=Entry(customer_frame,width=25,font=('tajawal,12'),justify=CENTER)
           en_name.place(x=10,y=40)
           l_total=Label(customer_frame,text=" السعر الكلي",bg='silver',fg='white')
           l_total.place(x=168,y=70)
           en_total=Entry(customer_frame,width=25,font=('tajawal,12'),justify=CENTER)
           en_total.place(x=10,y=90)
           l_date=Label(customer_frame,text="  تاريخ الشراء",bg='silver',fg='white')
           l_date.place(x=168,y=130)
           en_date=Entry(customer_frame,width=25,font=('tajawal,12'),justify=CENTER)
           en_date.place(x=10,y=150)
           buton_cler=Button(customer_frame,width=30,bg='#EDDBC0',cursor='hand2',text=" افراغ الحقول")
           buton_cler.place(x=12,y=300)
           buton_add=Button(customer_frame,width=30,bg='#EDDBC0',cursor='hand2',text="  حفظ الفاتورة",command=insertxlsx)
           buton_add.place(x=12,y=250)
           buton_delete=Button(customer_frame,width=30,bg='#EDDBC0',cursor='hand2',text="  حذف الفاتورة")
           buton_delete.place(x=12,y=200)
           buton_serch=Button(customer_frame,width=30,bg='#EDDBC0',cursor='hand2',text="   بحث مشتري")
           buton_serch.place(x=12,y=170)
           imglable.place(x=930,y=350)
           total=0
           for item in trevew.get_children():
               trevew.delete(item)
               
           for x in range(len(v_sp)):
               
               if int(v_sp[x].get())>0:
                  price=int(v_sp[x].get())* int(dict_price[x][1])
                  strval=(str(dict_price[x][1]),str(v_sp[x].get() ) ,str(price)   ) 
                  trevew.insert("",'end',iid=x,text=dict_price[x][0],values=strval)
                  total=price+total
           final=total    
           en_total.insert('1',str(final))  
           en_date.insert('1',str(date))
        def clearframe():
            for item in trevew.get_children():
                trevew.delete(item)
            en_name.delete('0',END)
            en_total.delete('0',END)
            en_date.delete('0',END)    
        win.geometry(self.wid_long)
        win.iconbitmap(self.ico)
        win.title(self.title)
        l_frame=Frame(win,bg='silver',width=600,height=550)
        l_frame.place(x=1,y=1)
        
        titl_frame=Label(l_frame,text='المواد والشراء والفاتور',font=('Tajawal 13'),fg='white',bg='#5F7161',width=70)
        titl_frame.place(x=0,y=0)
        
        
        imgbuton1=PhotoImage(file="img/2.png")
        imgbuton2=PhotoImage(file="img/3.png")
        buton1=Button(l_frame,width=100,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=200,image=imgbuton1,text=" قرطاسية ",compound=TOP)
        buton1.place(x=30,y=45)
        buton2=Button(l_frame,width=100,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=200,image=imgbuton2,text=" قميص ",compound=TOP)
        buton2.place(x=200,y=45)
        v_sp=[]
        v1=IntVar
        sp1=Spinbox(l_frame,fro_=0, to_= 5,textvariable=v1,width=10)
        sp1.place(x=30,y=245)
        v_sp.append(sp1)
        v2=IntVar
        sp2=Spinbox(l_frame,fro_=0, to_= 5,textvariable=v2,width=10)
        sp2.place(x=200,y=245)
        v_sp.append(sp2)
        r_frame=Frame(win,bg='gray',width=300,height=550)
        r_frame.place(x=610,y=1)
        trevew=ttk.Treeview(r_frame,selectmode='browse')
        trevew.place(x=1,y=1,width=300,height=300)
        trevew["columns"]=('1','2','3')
        trevew.column("#0",width=5,anchor='c')
        trevew.column("1",width=5,anchor='c')
        trevew.column("2",width=5,anchor='c')
        trevew.column("3",width=5,anchor='c')
        trevew.heading("#0",text='المواد',anchor='c')
        trevew.heading("1",text='السعر',anchor='c')
        trevew.heading("2",text='العدد',anchor='c')
        trevew.heading("3",text='المجموع',anchor='c')
        buton_sal=Button(l_frame,width=30,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=2,text=" شراء ",compound=TOP,command=bill)
        buton_sal.place(x=30,y=400)
        buton_newbill=Button(l_frame,width=30,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=2,text=" فاتورة جديدة",compound=TOP,command=clearframe)
        buton_newbill.place(x=200,y=400)
        imglablepath=PhotoImage(file="img/4.png")
        imglable=Label(win,image=imglablepath)
        win.mainloop()